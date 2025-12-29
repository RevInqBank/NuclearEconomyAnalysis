import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generate_income_statement(revenue, depreciation, cogs, sgna, tax_rate, interest):
    income_statements = []

    gross_profit = revenue - cogs
    ebitda = gross_profit - sgna
    ebit = ebitda - depreciation
    taxable_income = ebit - interest
    tax = taxable_income * tax_rate
    net_income = taxable_income - tax
    income_statements = np.vstack((revenue, cogs, gross_profit, sgna, ebitda, depreciation, ebit, interest, taxable_income, tax, net_income))
    return income_statements


# Example data for 10 years
revenue = np.array([1000, 1100, 1200, 1300, 1400, 1500, 1600, 1700, 1800, 1900])
depreciation = np.array([50, 55, 60, 65, 70, 75, 80, 85, 90, 95])
cogs = np.array([400, 420, 440, 460, 480, 500, 520, 540, 560, 580])
sgna = np.array([200, 210, 220, 230, 240, 250, 260, 270, 280, 290])
tax_rate = np.array([0.3, 0.3, 0.3, 0.3, 0.3, 0.3, 0.3, 0.3, 0.3, 0.3])
interest = np.array([20, 22, 24, 26, 28, 30, 32, 34, 36, 38])
        
# Generate the income statement
income_statements = generate_income_statement(revenue, depreciation, cogs, sgna, tax_rate, interest)


print(income_statements)

# Convert the income statement to a DataFrame with appropriate row and column names
row_names = ["Revenue", "COGS", "Gross Profit", "SG&A", "EBITDA", "Depreciation", "EBIT", "Interest", "Taxable Income", "Tax", "Net Income"]
column_names = [f"Year {i+1}" for i in range(income_statements.shape[1])]

income_statement_df = pd.DataFrame(income_statements, index=row_names, columns=column_names)

print(income_statement_df)

# Save the income statement DataFrame to an Excel file
output_file = "/Users/seungminkwak/income_statement.xlsx"
income_statement_df.to_excel(output_file, sheet_name="Income Statement", index=True)

print(f"Income statement saved to {output_file}")


# Apply formatting to the Excel file using openpyxl######################################################################################


# Load the Excel file
wb = load_workbook(output_file)
ws = wb["Income Statement"]


# Set font name for all cells
aptos_font = Font(name="Aptos Display")
for row in ws.iter_rows():
    for cell in row:
        cell.font = Font(name="Aptos Display", bold=cell.font.bold, color=cell.font.color)

# Style column index (header row)
for cell in ws[1]:
    cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    cell.font = Font(name="Aptos Display", bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="right", vertical="center")

# Style row index (first column)
for i, cell in enumerate(ws['A'][1:], start=1):  # Skip header
    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    bold_rows = ["Revenue", "Gross Profit", "EBIT", "Taxable Income", "Net Income"]
    if cell.value in bold_rows:
        cell.font = Font(name="Aptos Display", bold=True, color="000000")
    else:
        cell.font = Font(name="Aptos Display", bold=False, color="000000")
    cell.alignment = Alignment(horizontal="left", vertical="center")

# Adjust width of first column to fit text
max_length = max(len(str(cell.value)) for cell in ws['A'])
ws.column_dimensions['A'].width = max_length + 2

# Remove all borders from all cells
for row in ws.iter_rows():
    for cell in row:
        cell.border = Border()

# Add bold horizontal borders above specified rows
bold_border_rows = ["Revenue", "Gross Profit", "EBIT", "Taxable Income", "Net Income"]
for i, name in enumerate(ws['A'], start=1):
    if name.value in bold_border_rows:
        for cell in ws[i]:
            cell.border = Border(
                top=Side(style="medium"),
                left=Side(style=None),
                right=Side(style=None),
                bottom=Side(style=None)
            )

# Hide gridlines
ws.sheet_view.showGridLines = False

# Save the updated Excel file
wb.save(output_file)
print(f"Formatted income statement saved to {output_file}")