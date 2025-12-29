import numpy as np
import pandas as pd
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import numpy_financial as npf
import os


def _get_results_path():
    """output 폴더 안의 RESULTS.xlsx 경로 반환 (없으면 폴더 생성)."""
    current_directory = os.getcwd()
    output_dir = os.path.join(current_directory, "output")
    os.makedirs(output_dir, exist_ok=True)
    return os.path.join(output_dir, "RESULTS.xlsx")


def CASHFLOW(CFS):
    output_file = _get_results_path()
    sheet_name = "CASH FLOW STATEMENT"

    # 1) CFS DataFrame 저장
    with pd.ExcelWriter(output_file) as writer:
        CFS.to_excel(writer, sheet_name=sheet_name, index=True, header=False)

    # 2) 불러오기
    wb = load_workbook(output_file)
    ws = wb[sheet_name]

    # 3) 1행 스타일
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        cell.font = Font(name="Aptos Display", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 4) 첫 열 라벨 스타일
    bold_rows = ["GROSS PROFIT", "EBIT", "EBT (Taxable Income)", "NET INCOME", "CASH FLOW"]
    for i, cell in enumerate(ws['A'][2:], start=2):  # 2행부터
        if cell.value in bold_rows:
            cell.font = Font(name="Aptos Display", bold=True, color="000000")
        else:
            cell.font = Font(name="Aptos Display", bold=False, color="000000")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # 5) 첫 열 너비 자동 조정
    max_length = max(len(str(cell.value)) for cell in ws['A'] if cell.value is not None)
    ws.column_dimensions['A'].width = max_length + 2

    # 6) 모든 셀 테두리 제거
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border()

    # 7) 주요 행 위에 굵은 상단선
    for r, name_cell in enumerate(ws['A'], start=1):
        if name_cell.value in bold_rows:
            for cell in ws[r]:
                cell.border = Border(top=Side(style="medium"))

    # 8) 그리드 숨김
    ws.sheet_view.showGridLines = False

    wb.save(output_file)
    return output_file


def LCOE(CFS: pd.DataFrame, discountRate: float, electricityPrice, salesToRevenueRatio):
    def _row(row_name):
        if row_name not in CFS.index:
            raise KeyError(f"'{row_name}' 행이 없습니다.")
        s = pd.to_numeric(CFS.loc[row_name], errors="coerce").fillna(0.0).values.astype(float)
        return s

    def _to_array(x, n):
        if np.isscalar(x):
            return np.full(n, float(x))
        arr = np.asarray(x, dtype=float).flatten()
        if arr.size != n:
            raise ValueError(f"길이 불일치: 기대 {n}, 입력 {arr.size}")
        return arr

    def _npv(arr, r):
        t = np.arange(1, len(arr) + 1, dtype=float)
        return np.sum(arr / np.power(1.0 + r, t))

    # 데이터 추출
    rev = _row("REVENUE")
    capex = _row("CAPEX")
    interest = _row("INTEREST")
    cap_om = _row("Capital OM Cost")
    ann_om = _row("Annual OM Cost")
    fuel_fe = _row("FUEL (Front-end)")
    fuel_fe_is = _row("FUEL (Interim Storage)")

    n = len(rev)
    p = _to_array(electricityPrice, n)
    ratio = _to_array(salesToRevenueRatio, n)

    denom_series = - rev / (p * ratio)
    npv_denom = _npv(denom_series, discountRate)

    lcoe_con_num = _npv(capex + interest, discountRate)
    lcoe_om_num = _npv(cap_om + ann_om, discountRate)
    lcoe_fuel_num = _npv(fuel_fe, discountRate)
    lcoe_fuel_num_is = _npv(fuel_fe_is, discountRate)

    LCOE_CON = lcoe_con_num / npv_denom
    LCOE_OM = lcoe_om_num / npv_denom
    LCOE_FUEL = lcoe_fuel_num / npv_denom
    LCOE_FUEL_IS = lcoe_fuel_num_is / npv_denom
    LCOE_TOTAL = LCOE_CON + LCOE_OM + LCOE_FUEL

    # 엑셀 쓰기
    path = _get_results_path()
    sheet = "CASH FLOW STATEMENT"
    wb = load_workbook(path)
    ws = wb[sheet]

    ws["G23"].value = float(LCOE_CON)
    ws["I23"].value = float(LCOE_OM)
    ws["K23"].value = float(LCOE_FUEL)
    ws["D23"].value = float(LCOE_TOTAL)

    ws["D22"].value = "LCOE"
    ws["G22"].value = "Construction"
    ws["I22"].value = "O&M"
    ws["K22"].value = "Fuel"
    ws["E23"].value = "[$/MWh]"
    ws["F23"].value = "="
    ws["H23"].value = "+"
    ws["J23"].value = "+"

    center_bold = Alignment(horizontal="center", vertical="center")
    thick = Side(style="medium")
    bold_border = Border(left=thick, right=thick, top=thick, bottom=thick)

    result_cells = ["D23", "G23", "I23", "K23"]
    for addr in result_cells:
        cell = ws[addr]
        cell.alignment = center_bold
        cell.font = Font(bold=True, color="0070C0")
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        cell.border = bold_border

    label_cells = ["D22", "G22", "I22", "K22", "E23", "F23", "H23", "J23"]
    for addr in label_cells:
        c = ws[addr]
        c.alignment = center_bold
        c.font = Font(bold=True)

    wb.save(path)
    return LCOE_CON, LCOE_OM, LCOE_FUEL, LCOE_FUEL_IS, LCOE_TOTAL


def IRR(CFS):
    cash_flow = CFS.loc['CASH FLOW']
    cashflows = cash_flow.values
    try:
        irr = npf.irr(cashflows)
        if np.isnan(irr):
            print("IRR 계산 실패: 해가 존재하지 않음")
            return None

        file_path = _get_results_path()
        wb = load_workbook(file_path)
        ws = wb['CASH FLOW STATEMENT']

        ws['D25'].value = 'IRR'
        ws['D25'].font = Font(bold=True)
        ws['D25'].alignment = Alignment(horizontal='center')

        ws['D26'].value = irr
        ws['D26'].alignment = Alignment(horizontal='center')
        ws['D26'].font = Font(bold=True, color='FF0000')
        ws['D26'].fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
        ws['D26'].border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )
        ws['D26'].number_format = '0.00%'

        ws['E26'].value = '[%]'
        ws['E26'].font = Font(bold=True)
        ws['E26'].alignment = Alignment(horizontal='center')

        wb.save(file_path)
        return irr
    except Exception as e:
        print(f"IRR 계산 실패: {e}")
        return None


def BEP(CFS):
    try:
        cash_flow = CFS.loc['CASH FLOW']
        years = CFS.columns.tolist()
        cumulative_cash_flow = cash_flow.cumsum()

        bep_year = None
        for i in range(len(cumulative_cash_flow) - 1):
            if cumulative_cash_flow.iloc[i] < 0 and cumulative_cash_flow.iloc[i + 1] >= 0:
                fraction = abs(cumulative_cash_flow.iloc[i]) / (
                    abs(cumulative_cash_flow.iloc[i]) + cumulative_cash_flow.iloc[i + 1]
                )
                bep_year = years[i] + fraction * (years[i + 1] - years[i])
                break

        if bep_year is None:
            if cumulative_cash_flow.iloc[-1] < 0:
                print("BEP 없음: 전체 기간 손실")
                return None
            elif cumulative_cash_flow.iloc[0] >= 0:
                bep_year = float(years[0])

        if bep_year is not None:
            bep_rounded = round(bep_year, 2)
            file_path = _get_results_path()
            wb = load_workbook(file_path)
            ws = wb['CASH FLOW STATEMENT']

            ws['D28'].value = 'BEP (Break Even Point)'
            ws['D28'].font = Font(bold=True)
            ws['D28'].alignment = Alignment(horizontal='center')

            ws['D29'].value = bep_rounded
            ws['D29'].font = Font(bold=True, color='006400')
            ws['D29'].fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
            ws['D29'].border = Border(
                left=Side(style='medium'), right=Side(style='medium'),
                top=Side(style='medium'), bottom=Side(style='medium')
            )
            ws['D29'].alignment = Alignment(horizontal='center')
            ws['D29'].number_format = '0.00'

            ws['E29'].value = 'year'
            ws['E29'].font = Font(bold=True)
            ws['E29'].alignment = Alignment(horizontal='center')

            wb.save(file_path)
            return bep_rounded
        return None
    except Exception as e:
        print(f"BEP 계산 실패: {e}")
        return None


def CONSTRUCTION_COST(CFS):
    try:
        capex_total = CFS.loc['CAPEX'].sum()
        interest_total = CFS.loc['INTEREST'].sum()
        construction_total = capex_total + interest_total

        file_path = _get_results_path()
        wb = load_workbook(file_path)
        ws = wb['CASH FLOW STATEMENT']

        ws['D31'].value = 'Construction Cost'
        ws['D31'].font = Font(bold=True)
        ws['D31'].alignment = Alignment(horizontal='center')

        ws['G31'].value = 'CAPEX'
        ws['G31'].font = Font(bold=True)
        ws['G31'].alignment = Alignment(horizontal='center')

        ws['I31'].value = 'INTEREST'
        ws['I31'].font = Font(bold=True)
        ws['I31'].alignment = Alignment(horizontal='center')

        ws['D32'].value = abs(construction_total)
        ws['D32'].font = Font(bold=True, color='FF8C00')
        ws['D32'].fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
        ws['D32'].alignment = Alignment(horizontal='center')
        ws['D32'].number_format = '0.00'

        ws['E32'].value = '[$M]'
        ws['E32'].font = Font(bold=True)
        ws['E32'].alignment = Alignment(horizontal='center')

        ws['F32'].value = '='
        ws['F32'].font = Font(bold=True)
        ws['F32'].alignment = Alignment(horizontal='center')

        ws['G32'].value = abs(capex_total)
        ws['G32'].font = Font(bold=True, color='FF8C00')
        ws['G32'].fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
        ws['G32'].alignment = Alignment(horizontal='center')
        ws['G32'].number_format = '0.00'

        ws['H32'].value = '+'
        ws['H32'].font = Font(bold=True)
        ws['H32'].alignment = Alignment(horizontal='center')

        ws['I32'].value = abs(interest_total)
        ws['I32'].font = Font(bold=True, color='FF8C00')
        ws['I32'].fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
        ws['I32'].alignment = Alignment(horizontal='center')
        ws['I32'].number_format = '0.00'

        ws['J32'].value = '[$M]'
        ws['J32'].font = Font(bold=True)
        ws['J32'].alignment = Alignment(horizontal='center')

        wb.save(file_path)
        return abs(construction_total)
    except Exception as e:
        print(f"건설비 계산 실패: {e}")
        return None
